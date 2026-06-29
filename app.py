"""
Landing Page – Colombia Destinos
=================================
Backend Flask con base de datos SQLite para captura de leads.
"""

import os
import sqlite3
import io
from functools import wraps
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, Response, session, redirect, url_for, flash, g
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv


# ---------------------------------------------------------------------------
# Configuración de la aplicación
# ---------------------------------------------------------------------------
load_dotenv()
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "colombiadestinos-secret-key-change-in-production")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)  # La sesión expira en 30 minutos

# Configurar Limiter para evitar SPAM
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

@app.after_request
def add_security_headers(response):
    """Agrega cabeceras de seguridad a todas las respuestas."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

DATABASE = os.path.join(app.root_path, "leads.db")


# ---------------------------------------------------------------------------
# Utilidades de base de datos
# ---------------------------------------------------------------------------
def get_db():
    """Abre una conexión a la base de datos SQLite asociada al contexto global 'g'."""
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error):
    """Cierra la conexión de base de datos al final de cada petición."""
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """Crea las tablas 'leads' y 'packages' si no existen, y asegura sus columnas."""
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS leads (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre      TEXT    NOT NULL,
            telefono    TEXT    NOT NULL,
            destino     TEXT    NOT NULL,
            fecha_viaje TEXT    NOT NULL,
            fecha       TEXT    NOT NULL,
            contactado  INTEGER DEFAULT 0,
            oculto      INTEGER DEFAULT 0
        )
        """
    )
    cursor = conn.execute("PRAGMA table_info(leads)")
    columns = [row['name'] for row in cursor.fetchall()]
    if 'destino' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN destino TEXT DEFAULT ''")
    if 'fecha_viaje' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN fecha_viaje TEXT DEFAULT ''")
    if 'contactado' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN contactado INTEGER DEFAULT 0")
    if 'oculto' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN oculto INTEGER DEFAULT 0")
    if 'documento' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN documento TEXT DEFAULT ''")
    if 'correo' not in columns:
        conn.execute("ALTER TABLE leads ADD COLUMN correo TEXT DEFAULT ''")
    
    # Crear la tabla de paquetes si no existe
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS packages (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo          TEXT    NOT NULL,
            descripcion     TEXT    NOT NULL,
            clasificacion   TEXT    NOT NULL, -- 'nacional', 'internacional', 'pasadia'
            precio          INTEGER NOT NULL,
            precio_desde    INTEGER DEFAULT 0, -- 1 = "desde", 0 = "fijo"
            duracion_tipo   TEXT    NOT NULL, -- 'dias' o 'fecha'
            duracion_valor  TEXT    NOT NULL, -- ej. '30' o '2026-12-31'
            imagen          TEXT    NOT NULL, -- nombre de archivo de la imagen
            precios_variantes TEXT,          -- JSON string de variantes de precios
            fecha_creacion  TEXT    NOT NULL
        )
        """
    )
    conn.commit()


# Inicializar la base de datos al arrancar la aplicación
with app.app_context():
    init_db()


# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------

def get_active_packages():
    """Retorna los paquetes que no han expirado."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM packages ORDER BY id DESC").fetchall()
    active_packages = []
    now = datetime.now()
    for row in rows:
        pkg = dict(row)
        # Parsear variantes si existen
        if pkg.get('precios_variantes'):
            try:
                pkg['variantes'] = json.loads(pkg['precios_variantes'])
            except Exception:
                pkg['variantes'] = []
        else:
            pkg['variantes'] = []
            
        # Validar expiración
        fecha_creacion_dt = datetime.strptime(pkg['fecha_creacion'], "%Y-%m-%d %H:%M:%S")
        if pkg['duracion_tipo'] == 'dias':
            try:
                dias = int(pkg['duracion_valor'])
                fecha_expiracion = fecha_creacion_dt + timedelta(days=dias)
            except ValueError:
                fecha_expiracion = now + timedelta(days=1)
        elif pkg['duracion_tipo'] == 'fecha':
            try:
                fecha_expiracion = datetime.strptime(pkg['duracion_valor'], "%Y-%m-%d")
                fecha_expiracion = fecha_expiracion.replace(hour=23, minute=59, second=59)
            except ValueError:
                fecha_expiracion = now + timedelta(days=1)
        else:
            fecha_expiracion = now + timedelta(days=365)

        if now <= fecha_expiracion:
            pkg['dias_restantes'] = (fecha_expiracion - now).days
            pkg['fecha_expiracion_legible'] = fecha_expiracion.strftime("%d/%m/%Y")
            active_packages.append(pkg)
    return active_packages


def get_all_packages():
    """Retorna todos los paquetes, marcando si han expirado."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM packages ORDER BY id DESC").fetchall()
    packages = []
    now = datetime.now()
    for row in rows:
        pkg = dict(row)
        if pkg.get('precios_variantes'):
            try:
                pkg['variantes'] = json.loads(pkg['precios_variantes'])
            except Exception:
                pkg['variantes'] = []
        else:
            pkg['variantes'] = []
            
        fecha_creacion_dt = datetime.strptime(pkg['fecha_creacion'], "%Y-%m-%d %H:%M:%S")
        if pkg['duracion_tipo'] == 'dias':
            try:
                dias = int(pkg['duracion_valor'])
                fecha_expiracion = fecha_creacion_dt + timedelta(days=dias)
            except ValueError:
                fecha_expiracion = now
        elif pkg['duracion_tipo'] == 'fecha':
            try:
                fecha_expiracion = datetime.strptime(pkg['duracion_valor'], "%Y-%m-%d")
                fecha_expiracion = fecha_expiracion.replace(hour=23, minute=59, second=59)
            except ValueError:
                fecha_expiracion = now
        else:
            fecha_expiracion = now
            
        pkg['expirado'] = now > fecha_expiracion
        pkg['fecha_expiracion_legible'] = fecha_expiracion.strftime("%d/%m/%Y")
        packages.append(pkg)
    return packages



@app.route("/")
def index():
    """Renderiza la landing page principal."""
    return render_template("index.html")

@app.errorhandler(404)
def page_not_found(e):
    """Manejo de rutas no encontradas (error 404)."""
    return render_template("404.html"), 404


@app.errorhandler(429)
def ratelimit_handler(e):
    """Manejo amigable del error 429 (Límite de peticiones excedido)."""
    if request.path == "/guardar" or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({"status": "error", "message": e.description or "Demasiados intentos. Por favor espera un momento."}), 429
    flash(e.description or "Demasiados intentos. Por favor espera.", "error")
    return render_template("login.html"), 429


@app.route("/guardar", methods=["POST"])
@limiter.limit("5 per minute", error_message="Demasiados intentos. Por favor espera un momento.")
def guardar_lead():
    """
    Recibe los datos del formulario de contacto y los guarda
    en la base de datos.
    """
    nombre = request.form.get("nombre", "")[:100].strip()
    indicativo = request.form.get("indicativo", "+57")[:10].strip()
    telefono_raw = request.form.get("telefono", "")[:20].strip()
    documento = request.form.get("documento", "")[:50].strip()
    correo = request.form.get("correo", "")[:100].strip()
    destino = request.form.get("destino", "")[:100].strip()
    fecha_viaje = request.form.get("fecha_viaje", "")[:30].strip()

    if not nombre or not telefono_raw or not destino or not documento or not correo:
        return jsonify({"status": "error", "message": "Nombre, documento, correo, teléfono y destino son obligatorios."}), 400

    # Combinar indicativo + teléfono
    telefono = f"{indicativo} {telefono_raw}"
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO leads (nombre, telefono, documento, correo, destino, fecha_viaje, fecha) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (nombre, telefono, documento, correo, destino, fecha_viaje, fecha),
        )
        conn.commit()
        return jsonify({"status": "ok", "message": "¡Gracias! Un asesor experto se pondrá en contacto contigo pronto."})
    except Exception as e:
        return jsonify({"status": "error", "message": "Error al guardar los datos."}), 500


# ---------------------------------------------------------------------------
# Autenticación por Sesión
# ---------------------------------------------------------------------------

ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

def check_auth(username, password):
    """Verifica si las credenciales del administrador son correctas."""
    return username == ADMIN_USER and password == ADMIN_PASSWORD

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per 15 minutes", methods=["POST"], error_message="Demasiados intentos de inicio de sesión. Tu dispositivo ha sido bloqueado por 15 minutos.")
def login():
    """Página de inicio de sesión seguro."""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if check_auth(username, password):
            session.permanent = True  # Activa el timeout de 30 minutos
            session['logged_in'] = True
            return redirect(url_for('ver_leads'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Cierra la sesión y redirige al login."""
    session.pop('logged_in', None)
    return redirect(url_for('login'))


@app.route("/leads")
@requires_auth
def ver_leads():
    """Panel de administración para ver los leads capturados."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads WHERE oculto = 0 ORDER BY id DESC").fetchall()
    leads = []
    for row in rows:
        lead_dict = dict(row)
        try:
            dt = datetime.strptime(lead_dict['fecha'], "%Y-%m-%d %H:%M:%S")
            lead_dict['fecha_iso'] = dt.strftime("%Y-%m-%d")
            lead_dict['fecha_legible'] = dt.strftime("%d/%m/%Y")
            lead_dict['hora_legible'] = dt.strftime("%I:%M %p")
        except Exception:
            parts = lead_dict['fecha'].split(' ')
            lead_dict['fecha_iso'] = parts[0]
            lead_dict['fecha_legible'] = parts[0]
            lead_dict['hora_legible'] = parts[1] if len(parts) > 1 else ''
        leads.append(lead_dict)
    return render_template("admin.html", leads=leads)


@app.route("/leads/toggle_contactado/<int:lead_id>", methods=["POST"])
@requires_auth
def toggle_contactado(lead_id):
    """Alterna el estado de contactado de un lead en la base de datos."""
    conn = get_db()
    lead = conn.execute("SELECT contactado FROM leads WHERE id = ?", (lead_id,)).fetchone()
    if not lead:
        return jsonify({"status": "error", "message": "Lead no encontrado."}), 404
    
    nuevo_estado = 1 if lead['contactado'] == 0 else 0
    conn.execute("UPDATE leads SET contactado = ? WHERE id = ?", (nuevo_estado, lead_id))
    conn.commit()
    return jsonify({"status": "ok", "nuevo_estado": nuevo_estado})


@app.route("/leads/ocultar/<int:lead_id>", methods=["POST"])
@requires_auth
def ocultar_lead(lead_id):
    """Marca un lead como oculto (archivado) en la base de datos."""
    conn = get_db()
    lead = conn.execute("SELECT fecha, contactado FROM leads WHERE id = ?", (lead_id,)).fetchone()
    if not lead:
        return jsonify({"status": "error", "message": "Lead no encontrado."}), 404
    
    req_data = request.get_json() or {}
    forzar = req_data.get("force", False)

    if not forzar:
        if lead['contactado'] == 0:
            return jsonify({
                "status": "warning", 
                "message": "No puedes ocultar este cliente porque aún no ha sido marcado como **Contactado**."
            }), 400
            
        try:
            fecha_registro = datetime.strptime(lead['fecha'], "%Y-%m-%d %H:%M:%S")
        except Exception:
            parts = lead['fecha'].split(' ')
            fecha_registro = datetime.strptime(parts[0], "%Y-%m-%d")
            
        dias_transcurridos = (datetime.now() - fecha_registro).days
        if dias_transcurridos < 30:
            dias_restantes = 30 - dias_transcurridos
            return jsonify({
                "status": "warning", 
                "message": f"Este lead lleva registrado solo **{dias_transcurridos} días**. Para ocultarlo, deben transcurrir al menos **30 días** (faltan {dias_restantes} días)."
            }), 400
        
    conn.execute("UPDATE leads SET oculto = 1 WHERE id = ?", (lead_id,))
    conn.commit()
    return jsonify({"status": "ok", "message": "Lead ocultado (archivado) exitosamente."})


@app.route("/admin/packages/save", methods=["POST"])
@requires_auth
def save_package():
    """Crea o edita un paquete en la base de datos."""
    package_id = request.form.get("id")
    titulo = request.form.get("titulo", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    clasificacion = request.form.get("clasificacion", "nacional").strip()
    precio_raw = request.form.get("precio", "0").strip()
    precio_desde = 1 if request.form.get("precio_desde") else 0
    duracion_tipo = request.form.get("duracion_tipo", "dias").strip()
    duracion_valor = request.form.get("duracion_valor", "").strip()
    precios_variantes = request.form.get("precios_variantes", "[]").strip()

    if not titulo or not descripcion or not precio_raw or not duracion_valor:
        flash("Todos los campos obligatorios deben completarse.", "error")
        return redirect(url_for("ver_leads"))

    try:
        precio = int(precio_raw)
    except ValueError:
        flash("El precio debe ser un número válido.", "error")
        return redirect(url_for("ver_leads"))

    try:
        json.loads(precios_variantes)
    except ValueError:
        precios_variantes = "[]"

    imagen_filename = None
    file = request.files.get("imagen")
    conn = get_db()

    if package_id:
        row = conn.execute("SELECT imagen FROM packages WHERE id = ?", (package_id,)).fetchone()
        if not row:
            flash("El paquete a editar no existe.", "error")
            return redirect(url_for("ver_leads"))
        imagen_filename = row["imagen"]

    if file and file.filename != "":
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
            flash("Formato de imagen no válido. Use JPG, PNG, GIF o WEBP.", "error")
            return redirect(url_for("ver_leads"))
        
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_name = secure_filename(file.filename)
        new_filename = f"{timestamp}_{safe_name}"
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], new_filename)
        file.save(file_path)
        
        if package_id and imagen_filename:
            old_path = os.path.join(app.config["UPLOAD_FOLDER"], imagen_filename)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception:
                    pass
        imagen_filename = new_filename

    if not imagen_filename:
        flash("Debes subir un flyer para el paquete.", "error")
        return redirect(url_for("ver_leads"))

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if package_id:
        conn.execute(
            """
            UPDATE packages 
            SET titulo = ?, descripcion = ?, clasificacion = ?, precio = ?, precio_desde = ?, 
                duracion_tipo = ?, duracion_valor = ?, imagen = ?, precios_variantes = ?
            WHERE id = ?
            """,
            (titulo, descripcion, clasificacion, precio, precio_desde, duracion_tipo, duracion_valor, imagen_filename, precios_variantes, package_id)
        )
        flash("Paquete actualizado exitosamente.", "success")
    else:
        conn.execute(
            """
            INSERT INTO packages (titulo, descripcion, clasificacion, precio, precio_desde, duracion_tipo, duracion_valor, imagen, precios_variantes, fecha_creacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (titulo, descripcion, clasificacion, precio, precio_desde, duracion_tipo, duracion_valor, imagen_filename, precios_variantes, now_str)
        )
        flash("Paquete creado exitosamente.", "success")
    
    conn.commit()
    return redirect(url_for("ver_leads"))



@app.route("/leads/exportar")
@requires_auth
def exportar_leads():
    """Genera un archivo Excel (.xlsx) premium con todos los leads de la base de datos."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads ORDER BY id DESC").fetchall()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Colombia Destinos"
    
    # Activar cuadrícula visible
    ws.views.sheetView[0].showGridLines = True
    
    # Definir los encabezados
    headers = [
        "ID Lead", "Fecha de Registro", "Hora de Registro", 
        "Nombre del Lead", "Documento", "Correo", "Teléfono", "Destino de Interés", 
        "Fecha de Viaje", "Estado de Contacto", "Archivado (Oculto)"
    ]
    ws.append(headers)
    
    # Estilos del encabezado (Emerald Green premium para Colombia Destinos)
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # Emerald Green
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    font_data = Font(name="Segoe UI", size=10)
    border_thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    fill_contacted = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_pending = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    font_contacted = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    font_pending = Font(name="Segoe UI", size=10, bold=True, color="374151")

    for row_idx, row in enumerate(rows, start=2):
        lead_dict = dict(row)
        ws.row_dimensions[row_idx].height = 20
        
        fecha_legible = lead_dict['fecha']
        hora_legible = ""
        try:
            dt = datetime.strptime(lead_dict['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_legible = dt.strftime("%d/%m/%Y")
            hora_legible = dt.strftime("%I:%M %p")
        except Exception:
            parts = lead_dict['fecha'].split(' ')
            fecha_legible = parts[0]
            hora_legible = parts[1] if len(parts) > 1 else ''
            
        estado_texto = "Contactado" if lead_dict['contactado'] == 1 else "Sin contactar"
        archivado_texto = "Sí" if lead_dict['oculto'] == 1 else "No"
        
        ws.cell(row=row_idx, column=1, value=f"#{lead_dict['id']}").alignment = align_center
        ws.cell(row=row_idx, column=2, value=fecha_legible).alignment = align_center
        ws.cell(row=row_idx, column=3, value=hora_legible).alignment = align_center
        ws.cell(row=row_idx, column=4, value=lead_dict['nombre']).alignment = align_left
        ws.cell(row=row_idx, column=5, value=lead_dict.get('documento', '')).alignment = align_center
        ws.cell(row=row_idx, column=6, value=lead_dict.get('correo', '')).alignment = align_left
        ws.cell(row=row_idx, column=7, value=lead_dict['telefono']).alignment = align_center
        ws.cell(row=row_idx, column=8, value=lead_dict['destino']).alignment = align_left
        ws.cell(row=row_idx, column=9, value=lead_dict['fecha_viaje']).alignment = align_center
        
        cell_status = ws.cell(row=row_idx, column=10, value=estado_texto)
        cell_status.alignment = align_center
        if lead_dict['contactado'] == 1:
            cell_status.fill = fill_contacted
            cell_status.font = font_contacted
        else:
            cell_status.fill = fill_pending
            cell_status.font = font_pending
            
        ws.cell(row=row_idx, column=11, value=archivado_texto).alignment = align_center
            
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_num)
            if col_num != 10:
                c.font = font_data
            c.border = border_thin

    ws.auto_filter.ref = f"A1:K{len(rows) + 1}"
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f"leads_colombia_destinos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


@app.route("/robots.txt")
def robots():
    robots_txt = """User-agent: *
Allow: /

Disallow: /admin
Disallow: /login
Disallow: /exportar
Disallow: /logout

Sitemap: https://agenciacolombiadestinos.com/sitemap.xml
"""
    return Response(robots_txt, mimetype="text/plain")


@app.route("/sitemap.xml")
def sitemap():
    sitemap_xml = """<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
    <url>
        <loc>https://agenciacolombiadestinos.com/</loc>
        <changefreq>weekly</changefreq>
        <priority>1.0</priority>
    </url>
</urlset>
"""
    return Response(sitemap_xml, mimetype="application/xml")

# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, port=5001)
